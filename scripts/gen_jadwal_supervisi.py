"""
Generate Jadwal_Supervisi_Guru.xlsx (3 sheets: ICP, ICT, TCP) from the same corrected
schedule data used by the jadwalguruabbs web viewer.

Source: scratch_all_teachers.json (in this same scripts/ folder — regenerate it first via
gen_guru_data.py if the underlying export.xlsx has changed). Output goes to
../Jadwal_Supervisi_Guru.xlsx (jadwalguruabbs/ root, alongside index.html).

Rules confirmed with the user 2026-07-27:
- ICP sheet: teachers who teach neither ICT nor Quran. Columns: Nama Guru, Mapel,
  Jadwal 1 (Kelas/Hari/Jam), Jadwal 2 (Kelas/Hari/Jam).
- ICT sheet: same columns but Jadwal 1..5, only for ICT teachers.
- TCP sheet: Quran is team-taught, so the row unit is the CLASS, not the teacher (each
  class's Quran teacher-pair is consistent all week, verified empirically). Columns:
  Kelas, Guru, Jadwal 1, Jadwal 2 (Kelas/Hari/Jam each).
- "Jadwal N" selection: search forward through that teacher's/class's teaching days
  (in Monday..Saturday order, skipping non-teaching days) for the NEAREST day that has a
  genuine back-to-back double period (2 consecutive lessons, same class(es), with zero
  clock-time gap between them -- NOT just consecutive lesson numbers, since lesson 2-3 has
  a gender-specific break for boys' classes and lesson 6-7 always has the lunch break). If
  no remaining day has a double at all, fall back to the nearest remaining day's single
  period (this fallback applies to Jadwal 1 too, not just later slots). Each subsequent
  Jadwal starts searching from the day AFTER the previous pick.
"""
import json
import os
from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BASE = os.path.dirname(os.path.abspath(__file__))
DATA_PATH = os.path.join(BASE, "scratch_all_teachers.json")
OUT_PATH = os.path.join(BASE, "..", "Jadwal_Supervisi_Guru.xlsx")

DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
DAY_ID = {"Monday": "Senin", "Tuesday": "Selasa", "Wednesday": "Rabu", "Thursday": "Kamis",
          "Friday": "Jumat", "Saturday": "Sabtu"}
BOYS_LETTERS = "ABC"

WEEKDAY_BASE = {1: ("07.30", "08.10"), 2: ("08.10", "08.50"),
                4: ("09.50", "10.30"), 5: ("10.30", "11.10"), 6: ("11.10", "11.50"),
                7: ("13.00", "13.40"), 8: ("13.40", "14.20"), 9: ("14.20", "15.00")}
WEEKDAY_L3 = {"boys": ("09.10", "09.50"), "girls": ("08.50", "09.30")}
FRIDAY_BASE = {1: ("07.30", "08.10"), 2: ("08.10", "08.50"),
               4: ("09.50", "10.30"), 5: ("10.30", "11.10"),
               7: ("13.00", "13.40"), 8: ("13.40", "14.20"), 9: ("14.20", "15.00")}
FRIDAY_L3 = WEEKDAY_L3
SATURDAY_BASE = {1: ("07.15", "07.45"), 2: ("07.45", "08.15"),
                  4: ("09.00", "09.30"), 5: ("09.30", "10.00"), 6: ("10.00", "10.30")}
SATURDAY_L3 = {"boys": ("08.15", "08.45"), "girls": ("08.30", "09.00")}
DAY_SCHEDULE = {
    "Monday": (WEEKDAY_BASE, WEEKDAY_L3), "Tuesday": (WEEKDAY_BASE, WEEKDAY_L3),
    "Wednesday": (WEEKDAY_BASE, WEEKDAY_L3), "Thursday": (WEEKDAY_BASE, WEEKDAY_L3),
    "Friday": (FRIDAY_BASE, FRIDAY_L3), "Saturday": (SATURDAY_BASE, SATURDAY_L3),
}
VALID_LESSONS_BY_DAY = {
    "Monday": list(range(1, 10)), "Tuesday": list(range(1, 10)),
    "Wednesday": list(range(1, 10)), "Thursday": list(range(1, 10)),
    "Friday": [1, 2, 3, 4, 5, 7, 8, 9], "Saturday": list(range(1, 7)),
}


def gender_of(cls):
    return "boys" if cls[1:].strip()[0] in BOYS_LETTERS else "girls"


def time_of(day, lesson, gender):
    base, l3 = DAY_SCHEDULE[day]
    if lesson == 3:
        return l3[gender]
    return base[lesson]


def lesson_classes_by_day(cells, subject_filter=None):
    """day -> {lesson: sorted deduped class list} for lessons this teacher/class has content,
    optionally restricted to a single subject (used for the TCP per-class Quran view)."""
    result = {}
    for day in DAYS:
        day_map = {}
        for lesson in VALID_LESSONS_BY_DAY[day]:
            entries = cells.get(f"{day}|{lesson}")
            if not entries:
                continue
            if subject_filter:
                entries = [e for e in entries if e["subject"] == subject_filter]
                if not entries:
                    continue
            classes = sorted(set(e["class"] for e in entries))
            day_map[lesson] = classes
        result[day] = day_map
    return result


def find_doubles_and_singles(day_lesson_classes):
    """Returns (doubles_by_day, singles_by_day). A double is 2 lessons, same class-set,
    with zero clock-time gap (checked via time_of, not just adjacent lesson numbers --
    lesson 2-3 has a boys-only break, lesson 6-7 always has the lunch break)."""
    doubles_by_day = {}
    singles_by_day = {}
    for day in DAYS:
        day_map = day_lesson_classes[day]
        lessons_sorted = sorted(day_map.keys())
        doubles = []
        for l in lessons_sorted:
            if l + 1 in day_map and day_map[l] == day_map[l + 1]:
                gender = gender_of(day_map[l][0])
                t1 = time_of(day, l, gender)
                t2 = time_of(day, l + 1, gender)
                if t1[1] == t2[0]:
                    doubles.append((l, l + 1, day_map[l], t1[0], t2[1]))
        singles = []
        for l in lessons_sorted:
            gender = gender_of(day_map[l][0])
            t = time_of(day, l, gender)
            singles.append((l, day_map[l], t[0], t[1]))
        doubles_by_day[day] = doubles
        singles_by_day[day] = singles
    return doubles_by_day, singles_by_day


def pick_jadwal_slots(day_lesson_classes, n_slots):
    doubles_by_day, singles_by_day = find_doubles_and_singles(day_lesson_classes)
    teaching_days = [d for d in DAYS if singles_by_day[d]]
    picks = []
    start_idx = 0
    for _ in range(n_slots):
        remaining = teaching_days[start_idx:]
        chosen_day = None
        chosen_is_double = False
        for day in remaining:
            if doubles_by_day[day]:
                chosen_day = day
                chosen_is_double = True
                break
        if chosen_day is None and remaining:
            chosen_day = remaining[0]
            chosen_is_double = False
        if chosen_day is None:
            picks.append(None)
            continue
        if chosen_is_double:
            l1, l2, classes, start, end = doubles_by_day[chosen_day][0]
            jam_label = f"Jam ke-{l1}-{l2} ({start}-{end})"
        else:
            l, classes, start, end = singles_by_day[chosen_day][0]
            jam_label = f"Jam ke-{l} ({start}-{end})"
        picks.append({
            "kelas": ", ".join(classes),
            "hari": DAY_ID[chosen_day],
            "jam": jam_label,
        })
        start_idx = teaching_days.index(chosen_day) + 1
    return picks


# ---------- load data ----------
with open(DATA_PATH, encoding="utf-8") as f:
    data = json.load(f)
teachers = data["teachers"]
order = data["order"]

ict_nicks = [n for n in order if "ICT" in teachers[n]["stats"]["subjects"]]
quran_nicks = [n for n in order if "Quran" in teachers[n]["stats"]["subjects"]]
icp_nicks = [n for n in order if n not in ict_nicks and n not in quran_nicks]

# ---------- ICP / ICT rows (per teacher) ----------
# Per-teacher subject exclusion confirmed with the user 2026-07-27: these 2 teachers teach a
# second subject alongside their main one, and that second subject should never be offered as
# a supervision slot for them (their Mapel column still lists both subjects -- only the Jadwal
# picks are restricted).
EXCLUDE_SUBJECT_FOR_TEACHER = {
    "Mr Febri": "Sc",     # ICT sheet -- only ICT sessions are eligible, not his Sc sessions
    "Ms Fatma": "Math",   # ICP sheet -- only Sc sessions are eligible, not her Math sessions
}


def build_teacher_rows(nicks, n_slots):
    rows = []
    for nick in nicks:
        t = teachers[nick]
        excluded_subject = EXCLUDE_SUBJECT_FOR_TEACHER.get(nick)
        # exclude Leadership entries -- those are staff meeting blocks, not a class a
        # supervisor should observe, but they can otherwise look like an ordinary
        # double-period "class" (e.g. Leadership 7 Monday jam 5-6) to the picker below
        filtered_cells = {
            key: [e for e in entries if not e["class"].startswith("Leadership ")
                  and e["subject"] != excluded_subject]
            for key, entries in t["cells"].items()
        }
        filtered_cells = {k: v for k, v in filtered_cells.items() if v}
        dlc = lesson_classes_by_day(filtered_cells)
        picks = pick_jadwal_slots(dlc, n_slots)
        rows.append({
            "nama": t["fullname"],
            "mapel": ", ".join(t["stats"]["subjects"]),
            "jadwal": picks,
        })
    return rows


icp_rows = build_teacher_rows(icp_nicks, 2)
ict_rows = build_teacher_rows(ict_nicks, 5)

# ---------- TCP rows (per class) ----------
wsc_classes = []
for nick in quran_nicks:
    for entries in teachers[nick]["cells"].values():
        for e in entries:
            if e["subject"] == "Quran":
                wsc_classes.append(e["class"])
all_quran_classes = sorted(set(wsc_classes))


def class_sort_key(cls):
    m = cls.split(" ")[0]
    grade = int(m[0])
    letter = m[1]
    track = cls[len(m):].strip()
    return (grade, letter, track)


all_quran_classes.sort(key=class_sort_key)

tcp_rows = []
for cls in all_quran_classes:
    # gather this class's Quran-only cells across whichever teacher(s) teach it, and the
    # fixed teacher combo for the "Guru" column (verified constant across the week per class)
    combined_cells = defaultdict(list)
    teacher_combo = set()
    for nick in quran_nicks:
        t = teachers[nick]
        for key, entries in t["cells"].items():
            for e in entries:
                if e["subject"] == "Quran" and e["class"] == cls:
                    combined_cells[key].append(e)
                    teacher_combo.add(nick)
    dlc = lesson_classes_by_day(combined_cells)
    picks = pick_jadwal_slots(dlc, 2)
    tcp_rows.append({
        "kelas": cls,
        "guru": ", ".join(sorted(teacher_combo)),
        "jadwal": picks,
    })

# ---------- write Excel ----------
wb = openpyxl.Workbook()
wb.remove(wb.active)

HEADER_FILL = PatternFill("solid", fgColor="1F5C4C")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def write_sheet(ws, lead_cols, n_slots, rows, lead_keys):
    headers = list(lead_cols)
    for i in range(1, n_slots + 1):
        headers += [f"Jadwal {i} - Kelas", f"Jadwal {i} - Hari", f"Jadwal {i} - Jam"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER

    for r, row in enumerate(rows, start=2):
        col = 1
        for key in lead_keys:
            c = ws.cell(row=r, column=col, value=row[key])
            c.alignment = LEFT
            c.border = BORDER
            col += 1
        for i in range(n_slots):
            pick = row["jadwal"][i] if i < len(row["jadwal"]) else None
            for field in ("kelas", "hari", "jam"):
                val = pick[field] if pick else ""
                c = ws.cell(row=r, column=col, value=val)
                c.alignment = CENTER if field != "jam" else LEFT
                c.border = BORDER
                col += 1

    ws.freeze_panes = "A2"
    for i, header in enumerate(headers, start=1):
        letter = openpyxl.utils.get_column_letter(i)
        if "Jam" in header:
            ws.column_dimensions[letter].width = 26
        elif "Nama" in header or header == "Guru":
            ws.column_dimensions[letter].width = 22
        elif "Mapel" in header:
            ws.column_dimensions[letter].width = 14
        else:
            ws.column_dimensions[letter].width = 16
    ws.row_dimensions[1].height = 28


ws_icp = wb.create_sheet("ICP")
write_sheet(ws_icp, ["Nama Guru", "Mapel"], 2, icp_rows, ["nama", "mapel"])

ws_ict = wb.create_sheet("ICT")
write_sheet(ws_ict, ["Nama Guru", "Mapel"], 5, ict_rows, ["nama", "mapel"])

ws_tcp = wb.create_sheet("TCP")
write_sheet(ws_tcp, ["Kelas", "Guru"], 2, tcp_rows, ["kelas", "guru"])

wb.save(OUT_PATH)
print("ICP:", len(icp_rows), "guru")
print("ICT:", len(ict_rows), "guru")
print("TCP:", len(tcp_rows), "kelas")
print("saved:", OUT_PATH)
