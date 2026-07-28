import json
import os
from collections import defaultdict

import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(BASE, "export.xlsx")  # not committed — copy the school's aSc export here, see README
OUT = os.path.join(BASE, "scratch_all_teachers.json")

SUBJECT_SHEETS = ["Sprt.", "Soc.", "Sc.", "Quran.", "Math.", "IFE.", "ICT.", "Eng.", "Cv.", "BI."]
DISPLAY = {"Sprt": "Sprt", "Soc": "Soc", "Sc": "Sc", "Quran": "Quran", "Math": "Math",
           "IFE": "IFE", "ICT": "ICT", "Eng": "Eng", "Cv": "Cv", "BI": "BI"}
BOYS_LETTERS = "ABC"
DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
DAY_ID = {"Monday": "Senin", "Tuesday": "Selasa", "Wednesday": "Rabu", "Thursday": "Kamis", "Friday": "Jumat", "Saturday": "Sabtu"}

LESSONS = list(range(1, 10))

# Fixed roster order carried over verbatim from the jadwalguruabbs site's existing DATA.order
# (subject-grouped: Math -> IPA -> ICT -> Indo -> English -> Social -> Civic -> PAI -> Quran, per
# its CHANGELOG.md) — NOT re-derived here, so this generator can't silently reshuffle the sidebar
# navigation order. If the teacher roster ever changes, update this list by hand.
TEACHER_ORDER = [
    "Ms Dila", "Ms Fatma", "Ms Robi", "Ms Rohmah", "Ms Wafda", "Mr Febri", "Ms Binti", "Ms Fitri",
    "Ms Nourma", "Mr Amar", "Mr Ifan", "Mr Jack", "Mr Musfiq", "Ms Leita", "Ms Qisthi", "Ms Isti",
    "Ms Putri", "Ms Syarah", "Mr Hang", "Mr Joe", "Mr Sharih", "Mr Arfian", "Mr Arvin", "Mr Pampam",
    "Ms Khoim", "Ms Yona", "Ms Zahro", "Ms Nia", "Ust Amin", "Ust Yoki", "Us Naya", "Us Usaidah",
    "Ust Andi", "Ust Fahmi", "Ust Ida", "Ust Rizqia",
]

# school-policy day length, confirmed with the user 2026-07-25 (the previously-generated
# buatdafa/Jadwal_Mingguan_v9.4.xlsx had this WRONG for Friday — it assumed a normal short
# lunch break, giving the two post-break periods the wrong clock times):
# - Mon-Thu: 9 periods, normal 40-min bell schedule, lunch 11.50-13.00.
# - Friday: periods 1-5 same as Mon-Thu, then an extended break 11.10-13.00 for Jumu'ah
#   prayer. Period 6 itself is entirely consumed by that break — it does not exist as a
#   teaching slot. The source data's own Lesson# column still numbers the 3 post-break
#   periods 6/7/8 (raw, pre-shift numbering), but the school refers to them as periods
#   7/8/9 (period 6 having been "used up" by the break) — confirmed 2026-07-26 with
#   concrete examples. So raw lesson N>=6 on Friday must be remapped to display period
#   N+1 before anything else (time lookup, JSON keys) touches it.
# - Saturday: a completely different 30-minute-period schedule starting 07.15, with its own
#   15-minute gender-offset break — and the gender roles are SWAPPED from the weekday rule:
#   girls break after period 2, boys break after period 3. Only 6 periods total, no lunch.
VALID_LESSONS_BY_DAY = {
    "Monday": set(range(1, 10)), "Tuesday": set(range(1, 10)),
    "Wednesday": set(range(1, 10)), "Thursday": set(range(1, 10)),
    "Friday": {1, 2, 3, 4, 5, 7, 8, 9},  # 6 is a gap (the Jumu'ah break), not a ceiling
    "Saturday": set(range(1, 7)),
}

FRIDAY_RAW_TO_DISPLAY = {6: 7, 7: 8, 8: 9}  # raw Lesson# from the source sheets -> real period


def remap_lesson(day, raw_lesson):
    if day == "Friday":
        return FRIDAY_RAW_TO_DISPLAY.get(raw_lesson, raw_lesson)
    return raw_lesson


WEEKDAY_BASE = {1: ("07.30", "08.10"), 2: ("08.10", "08.50"),
                4: ("09.50", "10.30"), 5: ("10.30", "11.10"), 6: ("11.10", "11.50"),
                7: ("13.00", "13.40"), 8: ("13.40", "14.20"), 9: ("14.20", "15.00")}
WEEKDAY_L3 = {"boys": ("09.10", "09.50"), "girls": ("08.50", "09.30")}

FRIDAY_BASE = {1: ("07.30", "08.10"), 2: ("08.10", "08.50"),
               4: ("09.50", "10.30"), 5: ("10.30", "11.10"),
               7: ("13.00", "13.40"), 8: ("13.40", "14.20"), 9: ("14.20", "15.00")}
FRIDAY_L3 = WEEKDAY_L3  # periods 1-5 unaffected by the Jumu'ah break, same gender-offset rule

SATURDAY_BASE = {1: ("07.15", "07.45"), 2: ("07.45", "08.15"),
                  4: ("09.00", "09.30"), 5: ("09.30", "10.00"), 6: ("10.00", "10.30")}
SATURDAY_L3 = {"boys": ("08.15", "08.45"), "girls": ("08.30", "09.00")}  # roles swapped vs weekday

DAY_SCHEDULE = {
    "Monday": (WEEKDAY_BASE, WEEKDAY_L3), "Tuesday": (WEEKDAY_BASE, WEEKDAY_L3),
    "Wednesday": (WEEKDAY_BASE, WEEKDAY_L3), "Thursday": (WEEKDAY_BASE, WEEKDAY_L3),
    "Friday": (FRIDAY_BASE, FRIDAY_L3), "Saturday": (SATURDAY_BASE, SATURDAY_L3),
}

# --- Leadership (jam kepemimpinan) ---
# 3 more sheets exist in the source besides the 10 core subjects above: LEADERSHIP 7 / LEADERSHIP 8
# / LEASDERSHIP 9 (the source's own typo for grade 9's sheet name — the Subjects sheet has the same
# typo, but its Short code "L9" is clean). Confirmed 2026-07-27: these are a once-a-week whole-grade
# double period (Grade 7 Monday P5-6, Grade 8 Tuesday P5-6, Grade 9 Wednesday P5-6), and the Contract
# column already includes these hours (verified: contract minus core-subject-only slot count equals
# exactly 2 x number of grades a teacher participates in, for all 24 participants, zero exceptions) —
# so adding these slots to the grid must NOT touch total_lessons, only fill in the previously-empty
# cells.
#
# Data-shape gotcha: unlike a normal subject sheet (each column = one class, e.g. Math has separate
# 7A/7B/... columns), each LEADERSHIP sheet mirrors the SAME full comma-joined class list identically
# across every one of that grade's columns (e.g. "LEADERSHIP 7." row for Monday P5 has 6 columns,
# each containing the literal string "7A,7B,7C ICT,7C TCP,7D,7E" — not one class per column). Reusing
# the normal per-class fan-out logic on this would multiply-count each teacher 36x for grade 7 alone
# (6 columns x 6 comma-tokens). So leadership is parsed as its own single whole-grade event per
# active (day, lesson): one entry per participating teacher, not one per mirrored class column.
LEADERSHIP_SHEETS = ["LEADERSHIP 7.", "LEADERSHIP 8.", "LEASDERSHIP 9."]
LEADERSHIP_CODE_OF_SHEET = {"LEADERSHIP 7.": "L7", "LEADERSHIP 8.": "L8", "LEASDERSHIP 9.": "L9"}
LEADERSHIP_LABEL = {"L7": "Leadership 7", "L8": "Leadership 8", "L9": "Leadership 9"}
LEADERSHIP_CLASS_MARKER = "__LEADERSHIP__"


def gender_of(cls):
    return "boys" if cls[1:].strip()[0] in BOYS_LETTERS else "girls"


def time_of(day, lesson, gender):
    base, l3 = DAY_SCHEDULE[day]
    if lesson == 3:
        return l3[gender]
    return base[lesson]


wb_src = openpyxl.load_workbook(SRC, data_only=True, read_only=True)

wsc = wb_src["Classes"]
all_classes = []
for row in wsc.iter_rows(min_row=2, max_row=200, max_col=1, values_only=True):
    if not row[0]:
        break
    all_classes.append(row[0])

wst = wb_src["Teachers"]
nickname = {}
fullname_of = {}
contract_of = {}
for row in wst.iter_rows(min_row=2, max_col=6, values_only=True):
    name, short, contract = row[1], row[2], row[5]
    if name:
        nickname[name.strip()] = short
        fullname_of[short] = name.strip()
        if contract is not None and str(contract).strip().isdigit():
            contract_of[short] = int(str(contract).strip())

wsl = wb_src["Lessons"]
teacher_of = {}
leadership_participants = defaultdict(set)  # code ("L7"/"L8"/"L9") -> set of nicknames
for row in wsl.iter_rows(min_row=2, max_col=8, values_only=True):
    teacher, cls_field, group, subj = row[0], row[1], row[2], row[3]
    if not subj or not cls_field or teacher == "Without teacher" or not teacher:
        continue
    names = [nickname.get(n.strip(), n.strip()) for n in str(teacher).split(",") if n.strip()]
    if subj in LEADERSHIP_LABEL:
        leadership_participants[subj].update(names)
        continue
    if subj not in DISPLAY:
        continue
    classes_in_row = [c.strip() for c in str(cls_field).split(",") if c.strip() in all_classes]
    for cls in classes_in_row:
        teacher_of.setdefault((cls, subj), set()).update(names)

per_teacher = defaultdict(lambda: defaultdict(list))
for sheetname in SUBJECT_SHEETS:
    key = sheetname.rstrip(".")
    ws = wb_src[sheetname]
    for row in ws.iter_rows(min_row=4, max_col=30):
        d = row[1].value
        raw_lesson = row[2].value
        if d is None or raw_lesson is None or str(d).strip() == "":
            continue
        lesson = remap_lesson(d, int(raw_lesson))
        for cell in row[3:]:
            v = cell.value
            if not v:
                continue
            for cls in [x.strip() for x in str(v).split(",")]:
                if cls not in all_classes:
                    continue
                for t in teacher_of.get((cls, key), []):
                    # a teacher can be assigned to more than one class in the same
                    # day+lesson slot (team-teaching / parallel groups, e.g. Quran) —
                    # keep every occurrence instead of letting later ones overwrite earlier ones
                    per_teacher[t][(d, str(lesson))].append((cls, key))

# leadership: one whole-grade event per active (day, lesson) — see comment block above
leadership_slots = defaultdict(set)  # code -> set of (day, display_lesson)
for sheetname in LEADERSHIP_SHEETS:
    code = LEADERSHIP_CODE_OF_SHEET[sheetname]
    ws = wb_src[sheetname]
    for row in ws.iter_rows(min_row=4, max_col=30):
        d = row[1].value
        raw_lesson = row[2].value
        if d is None or raw_lesson is None or str(d).strip() == "":
            continue
        if not any(cell.value for cell in row[3:]):
            continue
        lesson = remap_lesson(d, int(raw_lesson))
        leadership_slots[code].add((d, lesson))

for code, slots in leadership_slots.items():
    for t in leadership_participants.get(code, []):
        for (day, lesson) in slots:
            per_teacher[t][(day, str(lesson))].append((LEADERSHIP_CLASS_MARKER, code))

# --- Whole-class-group "mapel tanpa guru" (no-teacher subjects) for the per-class view only ---
# Confirmed 2026-07-28: 4 more sheets besides LEADERSHIP have no teacher at all (matches
# "Without teacher" in the Lessons sheet), so they're invisible in the per-teacher grid (nothing
# to attribute) but should still show as occupied time in the per-class view:
#   HOMEROOM TEACHER -> "HT": Monday P1 + Friday P1, all 20 classes
#   SCOUT             -> "Scout": Saturday P5-6, grade 7+8 only (13 classes)
#   SENI BUDAYA KESENIAN -> "SBK": Friday P8-9 (raw 7-8, remapped — verified with remap_lesson,
#     don't trust the sheet's own raw Lesson# column directly, same trap as Friday elsewhere)
#   SELF DEVELOPMENT  -> "SD": Thursday P8-9, grade 7+8 only
# Same mirrored-duplicate data shape as LEADERSHIP (every column in the row repeats the identical
# full class-list string) — take the class list from just the first non-empty cell, not a
# per-column fan-out (that would multiply-count exactly like the LEADERSHIP gotcha above).
CLASS_BLOCK_SHEETS = {
    "HOMEROOM TEACHER.": "HT",
    "SCOUT.": "Scout",
    "SENI BUDAYA KESENIAN.": "SBK",
    "SELF DEVELOPMENT.": "SD",
}

class_blocks = defaultdict(dict)  # class -> {"day|lesson": {"subject":.., "start":.., "end":..}}


def _first_class_list(row):
    first_val = next((c.value for c in row[3:] if c.value), None)
    if not first_val:
        return []
    return [x.strip() for x in str(first_val).split(",") if x.strip() in all_classes]


for sheetname, label in CLASS_BLOCK_SHEETS.items():
    ws = wb_src[sheetname]
    for row in ws.iter_rows(min_row=4, max_col=30):
        d = row[1].value
        raw_lesson = row[2].value
        if d is None or raw_lesson is None or str(d).strip() == "":
            continue
        classes_here = _first_class_list(row)
        if not classes_here:
            continue
        lesson = remap_lesson(d, int(raw_lesson))
        tm = time_of(d, lesson, "boys")  # none of these touch period 3, gender-neutral is safe
        for cls in classes_here:
            class_blocks[cls][f"{d}|{lesson}"] = {"subject": label, "start": tm[0], "end": tm[1]}

# Leadership again, this time keyed by real class (the pass above keys it by teacher with a fake
# "Leadership N" class marker instead) — independent second pass over the same sheets, for the
# per-class view.
for sheetname in LEADERSHIP_SHEETS:
    ws = wb_src[sheetname]
    for row in ws.iter_rows(min_row=4, max_col=30):
        d = row[1].value
        raw_lesson = row[2].value
        if d is None or raw_lesson is None or str(d).strip() == "":
            continue
        classes_here = _first_class_list(row)
        if not classes_here:
            continue
        lesson = remap_lesson(d, int(raw_lesson))
        tm = time_of(d, lesson, "boys")
        for cls in classes_here:
            class_blocks[cls][f"{d}|{lesson}"] = {"subject": "Leadership", "start": tm[0], "end": tm[1]}

order = [n for n in TEACHER_ORDER if n in per_teacher]
missing = [n for n in per_teacher if n not in TEACHER_ORDER]
if missing:
    # new teacher not in the hand-maintained order above — append at the end rather than
    # silently dropping them, but this should be fixed by updating TEACHER_ORDER by hand
    order += sorted(missing)

teachers = {}
for nick in order:
    sched = per_teacher[nick]
    cells = {}
    classes_seen, subjects_seen, days_seen = set(), set(), set()
    raw_slot_count = 0  # grid-derived count — kept for diagnostics only, NOT shown as "jam/mgg"
    for (day, lesson_s), entries in sched.items():
        lesson = int(lesson_s)
        cell_entries = []
        for cls, subj_key in entries:
            if cls == LEADERSHIP_CLASS_MARKER:
                tm = time_of(day, lesson, "boys")  # leadership is always P5/P6 (no gender split)
                cell_entries.append({
                    "gender": "mixed",
                    "start": tm[0], "end": tm[1],
                    "class": LEADERSHIP_LABEL[subj_key],
                    "subject": "",
                })
                raw_slot_count += 1
                continue
            tm = time_of(day, lesson, gender_of(cls))
            cell_entries.append({
                "gender": gender_of(cls),
                "start": tm[0], "end": tm[1],
                "class": cls, "subject": DISPLAY[subj_key],
            })
            classes_seen.add(cls)
            subjects_seen.add(DISPLAY[subj_key])
            raw_slot_count += 1
        cells[f"{day}|{lesson}"] = cell_entries
        days_seen.add(day)
    # "jam/mgg" comes from the official Contract column (Teachers sheet), not from counting
    # grid slots — merged/team-taught sessions (e.g. 2 classes combined into 1 lesson, or a
    # teacher co-listed across parallel Quran groups) make grid-derived counts ambiguous and
    # inconsistent with how the school itself tracks contracted hours. Leadership hours are
    # already folded into this Contract number (see comment block above) so they must NOT be
    # added again here.
    total_lessons = contract_of.get(nick, raw_slot_count)
    teachers[nick] = {
        "nickname": nick,
        "fullname": fullname_of.get(nick, nick),
        "cells": cells,
        "stats": {
            "total_lessons": total_lessons,
            "active_days": len(days_seen),
            "classes": sorted(classes_seen),
            "subjects": sorted(subjects_seen),
        },
    }

data = {
    "order": order,
    "teachers": teachers,
    "days": DAYS,
    "day_id": DAY_ID,
    "lessons": LESSONS,
    "valid_lessons_by_day": {d: sorted(v) for d, v in VALID_LESSONS_BY_DAY.items()},
    "class_blocks": {cls: blocks for cls, blocks in class_blocks.items()},
}

with open(OUT, "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False)

print("teachers:", len(order))
ranked = sorted(order, key=lambda t: teachers[t]["stats"]["total_lessons"])
print("lightest:", ranked[0], teachers[ranked[0]]["stats"]["total_lessons"])
print("median:", ranked[len(ranked) // 2], teachers[ranked[len(ranked) // 2]]["stats"]["total_lessons"])
print("busiest:", ranked[-1], teachers[ranked[-1]]["stats"]["total_lessons"])
print("saved:", OUT)
